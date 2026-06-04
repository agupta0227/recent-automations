# ============================================================
#  loaders.py
#  BU–CG Reconciliation Engine — Version 3.8.0
#
#  Responsible for:
#    - Reading Excel files efficiently (openpyxl read_only)
#    - Normalizing strings, dates, numerics
#    - Fuzzy BU name matching and conflict detection
#    - Validating columns.json against actual sheet width
#    - Loading and validating Top18_mapping.json (FEAT-01)
#    - Caching via @st.cache_data (keyed on file bytes)
#
#  v3.7.0 changes (FEAT-01):
#    - load_top18_mapping() — parses & validates the carrier-group
#      mapping JSON. Raises Top18MappingError on any validation
#      failure (empty lists, wrong types, empty file).
#    - normalize_for_match() — shared aggressive normalizer for
#      carrier-group ↔ NetSuite Entity matching (trim + strip
#      special chars + CAPS).
#
#  v3.5.0 changes (Test05):
#    - load_netsuite returns (clean_df, dropped_df)
#    - load_jpmc returns 5 DataFrames (mapped, unmapped + 3 drop buckets)
#
#  When NetSuite and JPMC move fully to Snowflake, replace
#  this file with snowflake_loaders.py — nothing else changes.
# ============================================================

import io
import json
import re

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string


# ─────────────────────────────────────────
# String / date / numeric normalizers
# ─────────────────────────────────────────

def load_config(file) -> dict:
    """Parse columns.json from an uploaded file object."""
    return json.load(file)


# ─────────────────────────────────────────
# Top 18 carrier-group mapping (FEAT-01)
# ─────────────────────────────────────────

class Top18MappingError(Exception):
    """
    Raised when Top18_mapping.json fails validation. The caller is
    expected to surface the message to the user and abort pre-processing.

    Validation rules (all must pass):
      - File parses as a JSON object (dict)
      - Every value is a list (not a string, not null)
      - Every list contains at least one non-empty Entity Line ID
        after trim+normalize
    """
    pass


def normalize_for_match(s: str) -> str:
    """
    Aggressive normalization used on BOTH sides of the carrier-group
    matching:
      1. Trim leading/trailing whitespace
      2. Strip ALL non-alphanumeric characters (punctuation, dashes,
         periods, slashes, ampersands, commas, internal spaces, etc.)
      3. Uppercase

    Examples:
      'Cigna Health, Inc.'  → 'CIGNAHEALTHINC'
      'CIGNA HEALTH INC'    → 'CIGNAHEALTHINC'
      'Aetna-Health/FL'     → 'AETNAHEALTHFL'

    More forgiving than `.upper()` alone — handles real-world cases
    where the JSON has nicely-formatted names while NetSuite has
    machine-formatted ones.
    """
    if s is None:
        return ""
    return re.sub(r"[^A-Za-z0-9]+", "", str(s)).upper()


def load_top18_mapping(file) -> dict[str, list[str]]:
    """
    Parse Top18_mapping.json. Returns a dict keyed by carrier-group
    DISPLAY name (preserving JSON casing for UI labels), with values
    being lists of NetSuite Entity Line IDs (raw, not yet normalized
    for matching — call normalize_for_match() at match time).

    Validation:
      - File must parse as JSON object (dict)
      - Each value must be a list
      - Each list must contain ≥ 1 non-empty Entity ID after trim
      - Empty strings inside lists are filtered out (defensive against
        trailing-comma typos); abort only if filtering leaves the list
        empty

    Raises:
        Top18MappingError — with a clear message describing the problem.
    """
    try:
        raw = json.load(file)
    except json.JSONDecodeError as e:
        raise Top18MappingError(
            f"Top18_mapping.json is not valid JSON: {e}"
        )

    if not isinstance(raw, dict):
        raise Top18MappingError(
            "Top18_mapping.json must be a JSON object (mapping carrier-group "
            "names to lists of Entity Line IDs). Got: "
            f"{type(raw).__name__}."
        )

    if not raw:
        raise Top18MappingError(
            "Top18_mapping.json is empty. At least one carrier group with "
            "one or more Entity Line IDs is required."
        )

    cleaned: dict[str, list[str]] = {}
    for key, value in raw.items():
        # Carrier-group names: preserve casing for display, but trim
        # whitespace so leading/trailing space typos don't create
        # near-duplicate keys.
        display_name = str(key).strip()
        if not display_name:
            raise Top18MappingError(
                "Top18_mapping.json contains an empty carrier-group name. "
                "Every entry must have a non-empty key."
            )

        if not isinstance(value, list):
            raise Top18MappingError(
                f"Carrier group '{display_name}' must map to a JSON list of "
                f"Entity Line IDs. Got: {type(value).__name__}. "
                f"Example: \"{display_name}\": [\"ENTITY1\", \"ENTITY2\"]"
            )

        # Filter out empty/whitespace-only entries (defensive against
        # trailing commas or accidental "" entries in the JSON list)
        entities = [str(v).strip() for v in value if v is not None and str(v).strip()]

        if not entities:
            raise Top18MappingError(
                f"Carrier group '{display_name}' has no Entity Line IDs. "
                f"Every carrier group must have at least one Entity Line ID. "
                f"Either add IDs to '{display_name}' or remove the entry "
                f"from the JSON."
            )

        cleaned[display_name] = entities

    return cleaned


def normalize_str(series: pd.Series) -> pd.Series:
    """Strip whitespace and uppercase all string values."""
    return series.astype(str).str.strip().str.upper()


def normalize_fuzzy(series: pd.Series) -> pd.Series:
    """
    Strip ALL punctuation, spaces, dashes, periods, slashes for fuzzy key matching.
    e.g. 'Premier Life & Annuities, LLC' == 'PREMIER LIFE & ANNUITIES LLC'
    Used so minor formatting differences in BU names never cause missed matches.
    """
    return series.astype(str).str.upper().str.replace(r"[\s\.\,\-\/\&\']+", "", regex=True)


def normalize_date(series: pd.Series) -> pd.Series:
    """
    Parse dates tolerantly.
    format='mixed' handles inconsistent date formats across rows.
    dayfirst=False locks to MM/DD/YYYY for ambiguous dates (US convention).
    """
    return pd.to_datetime(series, errors="coerce", format="mixed", dayfirst=False)


def clean_numeric(series: pd.Series) -> pd.Series:
    """Coerce to float, fill non-parseable values with 0.0."""
    return pd.to_numeric(series, errors="coerce").fillna(0.0)


def add_year_month(df: pd.DataFrame, date_col: str) -> pd.DataFrame:
    """Add Year (Int64) and MonthName (string) columns derived from a date column."""
    df["Year"]      = df[date_col].dt.year.astype("Int64")
    df["MonthName"] = df[date_col].dt.strftime("%B")
    return df


# ─────────────────────────────────────────
# columns.json validation
# ─────────────────────────────────────────

def _sheet_max_column(file_bytes: bytes) -> int:
    """Return the number of columns (max_column) in the active sheet."""
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    max_col = ws.max_column or 0
    wb.close()
    return max_col


def validate_columns_config(
    cfg: dict,
    ns_bytes: bytes,
    jp_bytes: bytes,
    bu_bytes: bytes,
    dep_bytes: bytes | None = None,
) -> list[str]:
    """
    Validate that every column letter in columns.json points to a real
    column in the corresponding sheet. Returns a list of human-readable
    error messages — empty list means all good.

    Catches the silent-failure case where a wrong column letter (e.g.
    'BU': 'Z' on a 15-column sheet) would have produced None values
    throughout the loaded DataFrame with no warning.

    dep_bytes is optional — only validated when the Deposits file is
    uploaded. Skipped silently when None.
    """
    errors: list[str] = []

    sheets: dict[str, tuple[bytes, dict]] = {
        "netsuite":   (ns_bytes, cfg.get("netsuite", {})),
        "jpmc":       (jp_bytes, cfg.get("jpmc", {})),
        "bu_mapping": (bu_bytes, cfg.get("bu_mapping", {})),
    }
    if dep_bytes is not None:
        sheets["deposits"] = (dep_bytes, cfg.get("deposits", {}))

    for sheet_name, (file_bytes, sheet_cfg) in sheets.items():
        if not sheet_cfg:
            errors.append(f"columns.json is missing the '{sheet_name}' section.")
            continue

        try:
            # Detect CSV vs Excel. Deposits is typically CSV; others are xlsx.
            # CSV max-col = number of fields in the header row.
            is_csv = (
                sheet_name == "deposits" and
                b"\x50\x4b" not in file_bytes[:4]  # not a zip/xlsx header
            )
            if is_csv:
                # Read the first line to count columns
                first_line = file_bytes.decode("utf-8-sig", errors="replace").split("\n")[0]
                import csv as _csv
                max_col = len(next(_csv.reader([first_line])))
            else:
                max_col = _sheet_max_column(file_bytes)
        except Exception as e:
            errors.append(f"Could not open '{sheet_name}' file to validate columns: {e}")
            continue

        for field_name, letter in sheet_cfg.items():
            try:
                idx = column_index_from_string(str(letter).upper())
            except Exception:
                errors.append(
                    f"columns.json → {sheet_name}.{field_name} = '{letter}' "
                    f"is not a valid Excel column letter."
                )
                continue

            if idx > max_col:
                errors.append(
                    f"columns.json → {sheet_name}.{field_name} = '{letter}' "
                    f"(column #{idx}) is out of range — sheet only has {max_col} columns."
                )

    return errors


# ─────────────────────────────────────────
# Fast Excel reader
# ─────────────────────────────────────────

def _read_file_columns(file_bytes: bytes, col_letters: list[str],
                       filename: str = "") -> dict[str, list]:
    """
    Read only the required columns from an .xlsx or .csv file.

    For .xlsx: uses openpyxl read_only=True — skips styles/formatting,
               5-10x faster than default on large files.
    For .csv:  reads via csv module, maps column letters to positional
               indices (A=0, B=1, etc.) same as Excel columns.

    col_letters — list of Excel-style letters e.g. ["B", "E", "J"]
    filename    — used to detect file type by extension (optional,
                  falls back to trying xlsx then csv)
    """
    import csv as _csv

    col_indices = {letter: column_index_from_string(letter) for letter in col_letters}
    max_col     = max(col_indices.values())
    ext         = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    result    = {letter: [] for letter in col_letters}

    # ── CSV path ──────────────────────────────────────────────────────────
    is_csv = ext == "csv" or (ext not in ("xlsx", "xls") and b"\x50\x4b" not in file_bytes[:4])
    if is_csv:
        text      = file_bytes.decode("utf-8-sig", errors="replace")
        reader    = _csv.reader(text.splitlines())
        first_row = True
        for row in reader:
            if first_row:
                first_row = False
                continue
            for letter, idx in col_indices.items():
                result[letter].append(row[idx - 1] if len(row) >= idx else None)
        return result

    # ── Excel path ────────────────────────────────────────────────────────
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active

    first_row = True
    for row in ws.iter_rows(max_col=max_col, values_only=True):
        if first_row:
            first_row = False
            continue
        for letter, idx in col_indices.items():
            result[letter].append(row[idx - 1] if len(row) >= idx else None)

    wb.close()
    return result


# Keep old name as alias so existing callers don't break
_read_xlsx_columns = _read_file_columns


# ─────────────────────────────────────────
# NetSuite loader
# ─────────────────────────────────────────

@st.cache_data(show_spinner=False)
def load_netsuite(file_bytes: bytes, cfg: dict) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Load and normalize the NetSuite Excel export.

    Steps:
      1. Read only required columns (BU, DATE, ENTITY, REMARKS, CREDIT)
      2. Normalize strings to UPPER, parse dates, coerce credits
      3. Capture rows missing BU or with unparseable dates as the "dropped"
         frame for the caller to surface (Test05)
      4. Add Year + MonthName columns to the survivors

    Returns:
        (clean_df, dropped_df) — caller is responsible for surfacing the
        dropped rows to the user (loaders are pure, no st.* calls).

    Cached on file_bytes — same file never re-parsed within a session.
    """
    c    = cfg["netsuite"]
    cols = _read_xlsx_columns(file_bytes, [c["BU"], c["DATE"], c["ENTITY"], c["REMARKS"], c["CREDIT"]])

    df = pd.DataFrame({
        "BU":      normalize_str(pd.Series(cols[c["BU"]])),
        "Date":    normalize_date(pd.Series(cols[c["DATE"]])),
        "Entity":  normalize_str(pd.Series(cols[c["ENTITY"]])),
        "Remarks": normalize_str(pd.Series(cols[c["REMARKS"]])),
        "Credit":  clean_numeric(pd.Series(cols[c["CREDIT"]])),
    })

    # ── Identify rows to drop, capture them, then keep the survivors ──────
    bad_bu_mask   = ~(df["BU"].notna() & (df["BU"] != "") & (df["BU"] != "NAN"))
    bad_date_mask = df["Date"].isna()
    drop_mask     = bad_bu_mask | bad_date_mask

    dropped_df = df[drop_mask].copy()
    if not dropped_df.empty:
        # Annotate why each row was dropped — useful when the user reviews
        dropped_df["Drop_Reason"] = (
            bad_bu_mask[drop_mask].map({True: "Missing BU"}).fillna("") + " " +
            bad_date_mask[drop_mask].map({True: "Unparseable Date"}).fillna("")
        ).str.strip()

    df = df[~drop_mask]
    df = add_year_month(df, "Date")

    return df.reset_index(drop=True), dropped_df.reset_index(drop=True)


# ─────────────────────────────────────────
# BU Mapping loader
# ─────────────────────────────────────────

@st.cache_data(show_spinner=False)
def load_bu_mapping(file_bytes: bytes, cfg: dict) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Load the BU mapping table and detect conflicts.

    Many-to-one is supported: multiple BU_FULL names can share one BU_SHORT.
    The fuzzy key (punctuation stripped) is used for deduplication so that
    'Osborn Insurance Group LLC' and 'Osborn Insurance Group, LLC' collapse
    to one row rather than being treated as different entries.

    Returns:
        clean_df    — deduplicated, conflict-free mapping ready for joining
        conflicts_df — rows where the same fuzzy BU_Full maps to 2+ short codes
                       (shown as UI warning — user must resolve in the Excel file)
    """
    c    = cfg["bu_mapping"]
    cols = _read_xlsx_columns(file_bytes, [c["BU"], c["BU_FULL"]])

    df = pd.DataFrame({
        "BU":      normalize_str(pd.Series(cols[c["BU"]])),
        "BU_Full": normalize_str(pd.Series(cols[c["BU_FULL"]])),
    })

    df = df[
        df["BU"].notna()      & (df["BU"] != "")      & (df["BU"] != "NAN") &
        df["BU_Full"].notna() & (df["BU_Full"] != "") & (df["BU_Full"] != "NAN")
    ]

    # Fuzzy key strips all punctuation — safe dupes (same code, different punctuation) collapse
    df["BU_Full_Fuzzy"] = normalize_fuzzy(df["BU_Full"])
    df = df.drop_duplicates(subset=["BU_Full_Fuzzy", "BU"])

    # Detect genuine conflicts: same fuzzy name → different short codes
    conflict_mask = df.duplicated(subset=["BU_Full_Fuzzy"], keep=False)
    return df[~conflict_mask].reset_index(drop=True), df[conflict_mask].reset_index(drop=True)


# ─────────────────────────────────────────
# JPMC loader
# ─────────────────────────────────────────

@st.cache_data(show_spinner=False)
def load_jpmc(
    file_bytes: bytes, bu_map: pd.DataFrame, cfg: dict
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    Load the JPMC bank statement and map BU full names to short codes.

    The merge uses the fuzzy key (punctuation stripped) on both sides so that
    minor differences in how JPMC spells an account name (spaces, commas, dashes)
    never cause a missed match against the BU mapping.

    Filters applied during normalization:
      1. Drop rows with missing BU_Full or unparseable Date (basic hygiene)
      2. Drop rows with Credit == $0.00 (no money moved → not a real deposit)
      3. Drop rows whose Remarks contain any JPMC_NOISE_TOKENS substring
         (e.g. CASH CONCENTRATION — internal sweeps, not customer deposits)

    Returns (Test05 — full dropped frames, not just counts):
        mapped_df       — rows that successfully matched a BU short code
        unmapped_df     — rows with no BU match
        dropped_basic_df — rows dropped due to missing BU_Full / unparseable date
        dropped_zero_df  — rows dropped because Credit == $0.00
        dropped_noise_df — rows dropped because Remarks matched a noise token

    This loader is pure — no st.info / st.warning calls — so cache hits
    never silently swallow status feedback. The caller (app.py) is
    responsible for displaying coverage stats, warnings, and the dropped
    rows in expanders.
    """
    from config import JPMC_NOISE_TOKENS

    c      = cfg["jpmc"]
    needed = [c["BU_FULL"], c["DATE"], c["CREDIT"], c["REMARKS"]]
    cols   = _read_xlsx_columns(file_bytes, needed)

    df = pd.DataFrame({
        "BU_Full": normalize_str(pd.Series(cols[c["BU_FULL"]])),
        "Date":    normalize_date(pd.Series(cols[c["DATE"]])),
        "Credit":  clean_numeric(pd.Series(cols[c["CREDIT"]])),
        "Remarks": normalize_str(pd.Series(cols[c["REMARKS"]])),
    })

    # ── Filter 1: basic hygiene ──────────────────────────────────────────
    bad_bu_mask   = ~(df["BU_Full"].notna() & (df["BU_Full"] != "") & (df["BU_Full"] != "NAN"))
    bad_date_mask = df["Date"].isna()
    drop_basic_mask = bad_bu_mask | bad_date_mask

    dropped_basic_df = df[drop_basic_mask].copy()
    if not dropped_basic_df.empty:
        dropped_basic_df["Drop_Reason"] = (
            bad_bu_mask[drop_basic_mask].map({True: "Missing BU_Full"}).fillna("") + " " +
            bad_date_mask[drop_basic_mask].map({True: "Unparseable Date"}).fillna("")
        ).str.strip()

    df = df[~drop_basic_mask]
    df = add_year_month(df, "Date")

    # ── Filter 2: zero-credit rows ───────────────────────────────────────
    # Round to 2dp first so float noise like 0.00000001 is treated as $0.
    zero_mask = df["Credit"].round(2) == 0.00
    dropped_zero_df = df[zero_mask].copy()
    df = df[~zero_mask]

    # ── Filter 3: noise tokens in Remarks ────────────────────────────────
    # Remarks is already UPPERCASE from normalize_str; tokens are uppercased
    # defensively so future config edits don't silently miss matches.
    if JPMC_NOISE_TOKENS:
        pattern = "|".join(re.escape(t.upper()) for t in JPMC_NOISE_TOKENS)
        noise_mask = df["Remarks"].str.contains(pattern, na=False, regex=True)
    else:
        noise_mask = pd.Series(False, index=df.index)
    dropped_noise_df = df[noise_mask].copy()
    df = df[~noise_mask]

    # Fuzzy key on JPMC side — matches mapping regardless of punctuation
    df["BU_Full_Fuzzy"] = normalize_fuzzy(df["BU_Full"])

    merged   = df.merge(bu_map[["BU", "BU_Full_Fuzzy"]], on="BU_Full_Fuzzy", how="left")
    unmapped = merged[merged["BU"].isna() | (merged["BU"] == "")].copy()
    mapped   = merged[merged["BU"].notna() & (merged["BU"] != "")].copy()

    return (
        mapped.reset_index(drop=True),
        unmapped.reset_index(drop=True),
        dropped_basic_df.reset_index(drop=True),
        dropped_zero_df.reset_index(drop=True),
        dropped_noise_df.reset_index(drop=True),
    )


# ─────────────────────────────────────────
# Deposits loader
# ─────────────────────────────────────────

@st.cache_data(show_spinner=False)
def load_deposits(
    file_bytes: bytes,
    cfg: dict,
    filename: str = "",
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    Load the Deposits table (Snowflake extract / manual CSV or Excel).
    Handles both .xlsx and .csv via _read_file_columns().
    """
    from config import ORIG_CO_TOKEN

    c      = cfg.get("deposits", {})
    needed = [c["BU"], c["DATE"], c["AMOUNT"], c["REMARKS"]]
    cols   = _read_file_columns(file_bytes, needed, filename=filename)

    df = pd.DataFrame({
        "BU":      normalize_str(pd.Series(cols[c["BU"]])),
        "Date":    normalize_date(pd.Series(cols[c["DATE"]])),
        "Amount":  clean_numeric(pd.Series(cols[c["AMOUNT"]])),
        "Remarks": normalize_str(pd.Series(cols[c["REMARKS"]])),
    })

    # ── Drop rows with missing BU or unparseable date ─────────────────────
    basic_mask  = df["BU"].notna() & (df["BU"] != "") & (df["BU"] != "NAN") & df["Date"].notna()
    dropped_df  = df[~basic_mask].copy()
    df          = df[basic_mask].copy()

    df = add_year_month(df, "Date")

    # ── Filter to ORIG CO NAME rows only — same rule as NS and JPMC ───────
    origco_mask          = df["Remarks"].str.contains(ORIG_CO_TOKEN, na=False)
    dropped_no_origco_df = df[~origco_mask].copy()
    df                   = df[origco_mask].copy()

    return (
        df.reset_index(drop=True),
        dropped_df.reset_index(drop=True),
        dropped_no_origco_df.reset_index(drop=True),
    )
